#!/usr/bin/python3
import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
	row_id += 1

grade = []
row_id = 1
for row in ws:
	if row_id != 1:	
		grade.append(ws.cell(row = row_id, column = 7).value)
	row_id += 1
grade.sort(reverse=True)
length = len(grade)

grade_num = { 'A0' : 0, 'B0' : 0, 'C0' : 0 } 

row_id = 1
for row in ws:
	if row_id != 1 :
		score = ws.cell(row = row_id, column = 7).value
		index = grade.index(score)
		lastIndex = index + grade.count(score) - 1
		
		if lastIndex <= length * 0.3:
			ws.cell(row = row_id, column = 8).value = "A0"
			grade_num['A0'] += 1
		elif lastIndex <= length * 0.7:
			ws.cell(row = row_id, column = 8).value = "B0"
			grade_num['B0'] += 1
		else:
			ws.cell(row = row_id, column = 8).value = "C0"
			grade_num['C0'] += 1
	row_id += 1


row_id = 1
for row in ws:
	if row_id != 1:
		score = ws.cell(row = row_id, column = 7).value
		index = grade.index(score)
		lastIndex = index + grade.count(score)

		if lastIndex <= grade_num['A0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = "A+"
		elif lastIndex > grade_num['A0'] and lastIndex <= grade_num['A0'] + grade_num['B0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = "B+"
		elif lastIndex > grade_num['A0'] + grade_num['B0'] and lastIndex <= grade_num['A0'] + grade_num['B0'] + grade_num['C0'] * 0.5:
			ws.cell(row = row_id, column = 8).value = "C+"
	row_id += 1


wb.save("student.xlsx")
